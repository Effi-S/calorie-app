"""Tests for xlsx.py module."""
import unittest
import os
import tempfile
from unittest.mock import patch, MagicMock

from calorie_count.src.DB.food_db import FoodDB, Food
from calorie_count.src.DB.meal_entry_db import MealEntryDB, MealEntry
from calorie_count.src.utils import config
from calorie_count.src.utils.xlsx import save_to_excel, import_excel, DEFAULT_XLSX, FOOD_SHEET, MEALS_SHEET


class TestXlsx(unittest.TestCase):

    def setUp(self):
        """Set up test database and temporary Excel file."""
        self.test_db_path = config.set_db_path_test()
        self.temp_dir = tempfile.mkdtemp()
        self.test_xlsx_path = os.path.join(self.temp_dir, 'test_calorie.xlsx')
        
        # Set up test data
        with FoodDB() as fdb:
            fdb.add_food(Food('apple', 100, 0.5, 0.2, 10, 4, 0, 86))
            fdb.add_food(Food('banana', 100, 1.0, 0.3, 20, 15, 1, 75))
        
        with MealEntryDB() as mdb:
            with patch('calorie_count.src.DB.meal_entry_db.FoodDB.__enter__') as mock:
                mock.return_value = FoodDB()
                entry1 = MealEntry(name='apple', date='2022-12-15', portion=100)
                entry2 = MealEntry(name='banana', date='2022-12-16', portion=150)
                mdb.add_meal_entry(entry1)
                mdb.add_meal_entry(entry2)

    def tearDown(self):
        """Clean up temporary files."""
        if os.path.exists(self.test_xlsx_path):
            os.remove(self.test_xlsx_path)
        if os.path.exists(self.temp_dir):
            os.rmdir(self.temp_dir)
        
        # Clean up test database
        if hasattr(self, 'test_db_path') and os.path.exists(self.test_db_path):
            try:
                os.remove(self.test_db_path)
            except (OSError, PermissionError):
                pass  # Ignore if file is locked or already deleted

    def test_save_to_excel(self):
        """Test saving data to Excel file."""
        save_to_excel(self.test_xlsx_path)
        
        # Verify file was created
        self.assertTrue(os.path.exists(self.test_xlsx_path))
        
        # Verify file is not empty
        self.assertGreater(os.path.getsize(self.test_xlsx_path), 0)

    def test_save_to_excel_creates_sheets(self):
        """Test that Excel file contains correct sheets."""
        import openpyxl
        
        save_to_excel(self.test_xlsx_path)
        wb = openpyxl.load_workbook(self.test_xlsx_path)
        
        # Verify sheets exist
        self.assertIn(FOOD_SHEET, wb.sheetnames)
        self.assertIn(MEALS_SHEET, wb.sheetnames)

    def test_save_to_excel_food_data(self):
        """Test that food data is correctly saved to Excel."""
        import openpyxl
        
        save_to_excel(self.test_xlsx_path)
        wb = openpyxl.load_workbook(self.test_xlsx_path)
        ws = wb[FOOD_SHEET]
        
        # Check headers
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers, list(Food.columns()))
        
        # Check that we have food rows (header + foods)
        self.assertGreaterEqual(ws.max_row, 2)  # At least header + 1 food

    def test_save_to_excel_meal_data(self):
        """Test that meal entry data is correctly saved to Excel."""
        import openpyxl
        
        save_to_excel(self.test_xlsx_path)
        wb = openpyxl.load_workbook(self.test_xlsx_path)
        ws = wb[MEALS_SHEET]
        
        # Check headers
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers, list(MealEntry.columns()))
        
        # Check that we have meal rows (header + meals)
        self.assertGreaterEqual(ws.max_row, 2)  # At least header + 1 meal

    def test_import_excel(self):
        """Test importing data from Excel file."""
        import openpyxl
        
        # First save data to Excel
        save_to_excel(self.test_xlsx_path)
        
        # Clear the database
        with FoodDB() as fdb:
            for food_name in fdb.get_all_food_names():
                fdb.remove([food_name])
        
        # Import from Excel
        import_excel(self.test_xlsx_path)
        
        # Verify foods were imported
        with FoodDB() as fdb:
            food_names = fdb.get_all_food_names()
            self.assertIn('apple', food_names)
            self.assertIn('banana', food_names)

    def test_import_excel_invalid_sheet(self):
        """Test that importing with invalid sheet raises error."""
        import openpyxl
        
        # Create an Excel file with wrong headers
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = FOOD_SHEET
        ws.append(['Wrong', 'Headers', 'Here'])
        wb.save(self.test_xlsx_path)
        
        # Should raise AssertionError
        with self.assertRaises(AssertionError):
            import_excel(self.test_xlsx_path)

    def test_save_to_excel_empty_database(self):
        """Test saving empty database to Excel."""
        # Clear database
        with FoodDB() as fdb:
            for food_name in fdb.get_all_food_names():
                fdb.remove([food_name])
        
        empty_xlsx = os.path.join(self.temp_dir, 'empty.xlsx')
        save_to_excel(empty_xlsx)
        
        # File should still be created
        self.assertTrue(os.path.exists(empty_xlsx))
        
        # Clean up
        if os.path.exists(empty_xlsx):
            os.remove(empty_xlsx)


if __name__ == '__main__':
    unittest.main()
